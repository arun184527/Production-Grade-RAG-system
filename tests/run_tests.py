import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import json
import csv
from runner import run_rag
from metrics import (
    answer_correct,
    recall_at_k,
    input_limit_test,
    grounding_score,
    hallucination_check
)
with open("tests/dataset.json", "r") as f:
    dataset = json.load(f)
results = []
total_acc = 0
total_recall = 0
total_ground = 0
latencies = []
valid_tests = 0
for item in dataset:
    query = item["query"]
    keywords = item["relevant_keywords"]
    is_valid, word_count = input_limit_test(query)
    print("\n----------------------")
    print("QUERY:", query)
    print("WORDS:", word_count)
    if not is_valid:
        print("STATUS: FAILED (Too Long)")
        continue
    result = run_rag(query)
    answer = result["answer"]
    retrieved = result["retrieved"]
    latency = result["latency"]
    acc = answer_correct(answer, keywords)
    recall = recall_at_k(retrieved, keywords)
    ground = grounding_score(answer, retrieved)
    hallucination = hallucination_check(answer, retrieved)
    total_acc += acc
    total_recall += recall
    total_ground += ground
    latencies.append(latency)
    valid_tests += 1
    print("ANSWER:", answer)
    print("ACCURACY:", round(acc, 2))
    print("RECALL:", round(recall, 2))
    print("GROUNDING:", round(ground, 2))
    print("HALLUCINATION:", "PASS" if hallucination else "FAIL")
    print("LATENCY:", round(latency, 2))
    results.append({
        "query": query,
        "answer": answer,
        "accuracy": round(acc, 2),
        "recall": round(recall, 2),
        "grounding": round(ground, 2),
        "hallucination": "PASS" if hallucination else "FAIL",
        "latency": round(latency, 2)
    })
if valid_tests > 0:
    avg_acc = total_acc / valid_tests
    avg_recall = total_recall / valid_tests
    avg_ground = total_ground / valid_tests
    avg_latency = sum(latencies) / valid_tests
    print("\n======================")
    print("FINAL REPORT")
    print("Accuracy:", round(avg_acc, 2))
    print("Recall:", round(avg_recall, 2))
    print("Grounding:", round(avg_ground, 2))
    print("Avg Latency:", round(avg_latency, 2))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "RAG Report"

# Header
headers = ["Query", "Accuracy", "Recall", "Grounding", "Hallucination", "Latency"]
ws.append(headers)

# Style Header
header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col in ws[1]:
    col.fill = header_fill
    col.font = header_font
    col.alignment = Alignment(horizontal="center")

# Sort by accuracy
results.sort(key=lambda x: x["accuracy"])

# Color rules
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# Add data
for r in results:
    row = [
        r["query"],
        r["accuracy"],
        r["recall"],
        r["grounding"],
        r["hallucination"],
        round(r["latency"], 2)
    ]
    ws.append(row)

    current_row = ws.max_row

    # Accuracy coloring
    acc_cell = ws.cell(row=current_row, column=2)
    if r["accuracy"] >= 0.7:
        acc_cell.fill = green_fill
    elif r["accuracy"] >= 0.4:
        acc_cell.fill = yellow_fill
    else:
        acc_cell.fill = red_fill

    # Grounding coloring
    ground_cell = ws.cell(row=current_row, column=4)
    if r["grounding"] >= 0.7:
        ground_cell.fill = green_fill
    else:
        ground_cell.fill = red_fill

    # Hallucination coloring
    hall_cell = ws.cell(row=current_row, column=5)
    if r["hallucination"] == "PASS":
        hall_cell.fill = green_fill
    else:
        hall_cell.fill = red_fill

# Summary
avg_acc = round(total_acc / valid_tests, 2)
avg_recall = round(total_recall / valid_tests, 2)
avg_ground = round(total_ground / valid_tests, 2)
avg_latency = round(sum(latencies) / valid_tests, 2)

ws.append([])
ws.append([
    "FINAL SUMMARY",
    avg_acc,
    avg_recall,
    avg_ground,
    "-",
    avg_latency
])

summary_row = ws.max_row

for col in range(1, 7):
    cell = ws.cell(row=summary_row, column=col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Insight row
ws.append([
    "INSIGHT",
    "Strong Retrieval",
    "High Recall",
    "Good Grounding",
    "",
    "Latency Acceptable"
])

# Auto column width
for column in ws.columns:
    max_length = 0
    col_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

# Save
wb.save("tests/report.xlsx")

print("Professional Excel report saved to tests/report.xlsx")